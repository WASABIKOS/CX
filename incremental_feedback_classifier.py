import argparse
import hashlib
import json
import sqlite3
import urllib.request
from datetime import datetime, timezone
from pathlib import Path

import joblib
import openpyxl


TAXONOMY = [
    "Bajas y cancelaciones", "Compra, activación y portabilidad",
    "Cuenta, titularidad y app", "Facturación, pagos y crédito",
    "Fallas de internet residencial", "Fallas de servicio móvil",
    "Mudanza, instalación y visita", "Otros motivos y derivaciones",
    "Planes y cambios de plan", "Recargas y paquetes prepago",
    "Roaming internacional", "SIM y eSIM", "Saldo y consumo",
    "Sin motivo o abandono temprano", "TV y control remoto",
]

COMMENT_COLUMNS = [
    "rNPS - Overall Satisfaction comment",
    "Internet Additional Comments",
    "Phone Mobile Catchall Comment",
]


def now_iso():
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def normalize_text(value):
    return " ".join(str(value or "").split()).strip()


def comment_hash(text):
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def initialize_db(conn):
    conn.executescript("""
    PRAGMA journal_mode=WAL;
    CREATE TABLE IF NOT EXISTS responses (
        survey_key TEXT PRIMARY KEY,
        survey_id TEXT,
        cw_unique_id TEXT,
        response_date TEXT,
        survey_type TEXT,
        plan_type TEXT,
        product_segment TEXT,
        nps_class TEXT,
        score REAL,
        comment TEXT,
        comment_hash TEXT NOT NULL,
        source_file TEXT NOT NULL,
        source_mtime REAL NOT NULL,
        active INTEGER NOT NULL DEFAULT 1,
        first_seen_at TEXT NOT NULL,
        last_seen_at TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS classifications (
        survey_key TEXT NOT NULL,
        classifier TEXT NOT NULL,
        model_version TEXT NOT NULL,
        comment_hash TEXT NOT NULL,
        category TEXT,
        status TEXT NOT NULL,
        classified_at TEXT NOT NULL,
        PRIMARY KEY (survey_key, classifier)
    );
    CREATE TABLE IF NOT EXISTS classification_history (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        survey_key TEXT NOT NULL,
        classifier TEXT NOT NULL,
        model_version TEXT NOT NULL,
        comment_hash TEXT NOT NULL,
        category TEXT,
        status TEXT NOT NULL,
        classified_at TEXT NOT NULL
    );
    CREATE INDEX IF NOT EXISTS idx_responses_hash ON responses(comment_hash);
    CREATE INDEX IF NOT EXISTS idx_classifications_status ON classifications(classifier, status);
    """)


def extract_rows(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    iterator = ws.iter_rows(values_only=True)
    next(iterator, None)
    next(iterator, None)
    headers = list(next(iterator))
    positions = {name: i for i, name in enumerate(headers) if name}
    required = ["ID de encuesta", "CW - Unique ID", "Customer Response Date (EST)",
                "Survey Type", "Plan Type", "Broadband RGU",
                "Probabilidad de Recomendar", "Internet - Likelihood to Recommend",
                "Mobile - Likelihood to Recommend", *COMMENT_COLUMNS]
    missing = [name for name in required if name not in positions]
    if missing:
        raise ValueError(f"Columnas faltantes: {missing}")

    def get(row, name):
        value = row[positions[name]]
        return "" if value is None else value

    result = []
    for row in iterator:
        survey_id = normalize_text(get(row, "ID de encuesta")).removesuffix(".0")
        unique_id = normalize_text(get(row, "CW - Unique ID"))
        survey_key = survey_id or unique_id
        if not survey_key:
            continue
        comments = [normalize_text(get(row, name)) for name in COMMENT_COLUMNS]
        comment = " | ".join(value for value in comments if value)
        if not comment:
            continue
        survey_type = normalize_text(get(row, "Survey Type"))
        plan_type = normalize_text(get(row, "Plan Type"))
        broadband = normalize_text(get(row, "Broadband RGU"))
        score = None
        product = "No clasificado"
        if survey_type.lower() == "rnps":
            score = get(row, "Probabilidad de Recomendar")
            product = "rNPS / Relación"
        elif survey_type.lower() == "pnps" and plan_type.lower() == "servicio residencial" and broadband not in ("", "0"):
            score = get(row, "Internet - Likelihood to Recommend")
            product = "pNPS Internet"
        elif survey_type.lower() == "pnps" and "contrato" in plan_type.lower():
            score = get(row, "Mobile - Likelihood to Recommend")
            product = "pNPS Mobile - Contrato"
        elif survey_type.lower() == "pnps" and "prepago" in plan_type.lower():
            score = get(row, "Mobile - Likelihood to Recommend")
            product = "pNPS Mobile - Prepago"
        try:
            score = float(score)
            nps_class = "Promotor" if score >= 9 else "Neutro" if score >= 7 else "Detractor"
        except (TypeError, ValueError):
            score, nps_class = None, "Sin score"
        result.append({
            "survey_key": survey_key, "survey_id": survey_id,
            "cw_unique_id": unique_id,
            "response_date": normalize_text(get(row, "Customer Response Date (EST)"))[:10],
            "survey_type": survey_type, "plan_type": plan_type,
            "product_segment": product, "nps_class": nps_class,
            "score": score, "comment": comment, "comment_hash": comment_hash(comment),
        })
    wb.close()
    return result


def upsert_responses(conn, rows, source_path):
    timestamp = now_iso()
    source_mtime = source_path.stat().st_mtime
    conn.execute("UPDATE responses SET active=0")
    new_count = changed_count = unchanged_count = 0
    pending = []
    for row in rows:
        current = conn.execute(
            "SELECT comment_hash, first_seen_at FROM responses WHERE survey_key=?",
            (row["survey_key"],),
        ).fetchone()
        if current is None:
            new_count += 1
            pending.append(row)
            first_seen = timestamp
        elif current[0] != row["comment_hash"]:
            changed_count += 1
            pending.append(row)
            first_seen = current[1]
        else:
            unchanged_count += 1
            first_seen = current[1]
        conn.execute("""
            INSERT INTO responses VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            ON CONFLICT(survey_key) DO UPDATE SET
              survey_id=excluded.survey_id, cw_unique_id=excluded.cw_unique_id,
              response_date=excluded.response_date, survey_type=excluded.survey_type,
              plan_type=excluded.plan_type, product_segment=excluded.product_segment,
              nps_class=excluded.nps_class, score=excluded.score, comment=excluded.comment,
              comment_hash=excluded.comment_hash, source_file=excluded.source_file,
              source_mtime=excluded.source_mtime, active=1, last_seen_at=excluded.last_seen_at
        """, (
            row["survey_key"], row["survey_id"], row["cw_unique_id"], row["response_date"],
            row["survey_type"], row["plan_type"], row["product_segment"], row["nps_class"],
            row["score"], row["comment"], row["comment_hash"], str(source_path),
            source_mtime, 1, first_seen, timestamp,
        ))
    return pending, {"new": new_count, "changed": changed_count, "unchanged": unchanged_count}


def save_classification(conn, row, classifier, version, category, status="classified"):
    timestamp = now_iso()
    values = (row["survey_key"], classifier, version, row["comment_hash"], category, status, timestamp)
    conn.execute("INSERT INTO classification_history (survey_key,classifier,model_version,comment_hash,category,status,classified_at) VALUES (?,?,?,?,?,?,?)", values)
    conn.execute("""
        INSERT INTO classifications VALUES (?,?,?,?,?,?,?)
        ON CONFLICT(survey_key,classifier) DO UPDATE SET
          model_version=excluded.model_version, comment_hash=excluded.comment_hash,
          category=excluded.category, status=excluded.status, classified_at=excluded.classified_at
    """, values)


def classify_local(conn, rows, model_path):
    package = joblib.load(model_path)
    model = package["motivo_v1"]
    version = str(package.get("version", "unknown")) + "/motivo_v1"
    to_classify = []
    for row in rows:
        current = conn.execute("SELECT model_version,comment_hash FROM classifications WHERE survey_key=? AND classifier='local'", (row["survey_key"],)).fetchone()
        if not current or current[0] != version or current[1] != row["comment_hash"]:
            to_classify.append(row)
    if to_classify:
        matrix = model["vectorizer"].transform([row["comment"] for row in to_classify])
        labels = model["classifier"].predict(matrix)
        for row, label in zip(to_classify, labels):
            save_classification(conn, row, "local", version, str(label))
    return len(to_classify), version


def classify_ollama(conn, rows, model_name, limit):
    version = model_name
    candidates = []
    for row in rows:
        current = conn.execute("SELECT model_version,comment_hash,status FROM classifications WHERE survey_key=? AND classifier='ollama'", (row["survey_key"],)).fetchone()
        if not current or current[0] != version or current[1] != row["comment_hash"] or current[2] != "classified":
            candidates.append(row)
    candidates = candidates[:limit] if limit else []
    for row in candidates:
        prompt = ("Clasifica este comentario de telecomunicaciones en UNA categoría exacta. "
                  "Responde solamente el nombre exacto. Categorías: " + " | ".join(TAXONOMY) +
                  ". Comentario: " + row["comment"][:700])
        payload = json.dumps({"model": model_name, "prompt": prompt, "stream": False,
                              "options": {"temperature": 0, "num_predict": 40}}).encode()
        try:
            request = urllib.request.Request("http://127.0.0.1:11434/api/generate", data=payload,
                                             headers={"Content-Type": "application/json"})
            response = json.loads(urllib.request.urlopen(request, timeout=120).read())
            category = response.get("response", "").strip().strip("'\"")
            status = "classified" if category in TAXONOMY else "invalid_response"
            save_classification(conn, row, "ollama", version, category if status == "classified" else None, status)
        except Exception:
            save_classification(conn, row, "ollama", version, None, "error")
    return len(candidates), version


def main():
    parser = argparse.ArgumentParser(description="Clasificación incremental de feedbacks NPS")
    parser.add_argument("--input", default=r"C:\Claude\NPS\cwp_nps_responses_2026-08-14 10_16_11.xlsx")
    parser.add_argument("--db", default=r"outputs\medallia_cx_nps_2026-08-14\feedback_classifications.sqlite")
    parser.add_argument("--local-model", default=r"C:\Claude\MODEL\CAT_CX_MODEL_V4_2_SINTETICO.pkl")
    parser.add_argument("--ollama-model", default="qwen2.5-coder:1.5b")
    parser.add_argument("--ollama-limit", type=int, default=0,
                        help="Cantidad máxima de pendientes a enviar a Ollama; 0 no ejecuta Ollama")
    args = parser.parse_args()
    source_path, db_path = Path(args.input), Path(args.db)
    db_path.parent.mkdir(parents=True, exist_ok=True)
    rows = extract_rows(source_path)
    with sqlite3.connect(db_path) as conn:
        initialize_db(conn)
        pending, changes = upsert_responses(conn, rows, source_path)
        local_count, local_version = classify_local(conn, rows, Path(args.local_model))
        ollama_count, ollama_version = classify_ollama(conn, rows, args.ollama_model, args.ollama_limit)
        active = conn.execute("SELECT COUNT(*) FROM responses WHERE active=1").fetchone()[0]
        inactive = conn.execute("SELECT COUNT(*) FROM responses WHERE active=0").fetchone()[0]
        conn.commit()
    print(json.dumps({
        "source_rows_with_comment": len(rows), "active": active, "inactive": inactive,
        **changes, "local_classified_now": local_count, "local_version": local_version,
        "ollama_attempted_now": ollama_count, "ollama_version": ollama_version,
        "database": str(db_path.resolve()),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
