"""Create the review workbook used by the dashboard comment-export link."""

from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


HEADERS = (
    "Clave de encuesta", "Mes", "Segmento", "Clasificación NPS", "Puntaje",
    "Categoría automática", "Categoría final", "Fuente de categoría", "Comentario",
)
FIELDS = (
    "feedback_key", "month", "segment", "nps_class", "score", "category_auto",
    "category", "category_source", "feedback",
)


def write_categorized_comments_workbook(feedback_rows, output_path):
    """Write a filterable local review workbook without changing categories."""
    rows = list(feedback_rows)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Comentarios categorizados"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A5"

    sheet.merge_cells("A1:I1")
    sheet["A1"] = "Revisión de comentarios categorizados"
    sheet["A1"].fill = PatternFill("solid", fgColor="171717")
    sheet["A1"].font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 28

    sheet.merge_cells("A2:I2")
    sheet["A2"] = f"Base local: {len(rows):,} comentarios | Generado: {datetime.now():%Y-%m-%d %H:%M}"
    sheet["A2"].fill = PatternFill("solid", fgColor="F1F1F1")
    sheet["A2"].font = Font(name="Calibri", size=11, italic=True, color="4A4A4A")
    sheet["A2"].alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[2].height = 20

    for column, header in enumerate(HEADERS, start=1):
        cell = sheet.cell(4, column, header)
        cell.fill = PatternFill("solid", fgColor="E00032")
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[4].height = 28

    for row in rows:
        sheet.append([row.get(field, "") for field in FIELDS])

    sheet.auto_filter.ref = f"A4:I{len(rows) + 4}"
    for column, width in enumerate((42, 13, 30, 18, 10, 31, 31, 17, 92), start=1):
        sheet.column_dimensions[chr(64 + column)].width = width
    workbook.properties.title = "Comentarios categorizados CX NPS"
    workbook.properties.subject = "Revisión local de categorías y comentarios"
    workbook.save(output_path)
    return {"comments": len(rows), "columns": len(HEADERS)}
