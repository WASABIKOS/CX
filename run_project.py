"""Build the local CX NPS dataset and dashboard from a CWP survey export."""

import argparse
import csv
import hashlib
import json
import subprocess
import sys
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path

import openpyxl

from cx_taxonomy import TAXONOMY, TAXONOMY_VERSION, categorize


SPANISH_MONTHS = (
    "Ene", "Feb", "Mar", "Abr", "May", "Jun",
    "Jul", "Ago", "Sep", "Oct", "Nov", "Dic",
)


def report_timestamp():
    now = datetime.now()
    return f"{now.day:02d} {SPANISH_MONTHS[now.month - 1]} {now.year} · {now:%H:%M}"


COMMENT_COLUMNS = [
    "rNPS - Overall Satisfaction comment",
    "Internet Additional Comments",
    "Phone Mobile Catchall Comment",
]
NEEDED_COLUMNS = [
    "CW - Unique ID", "Unit", "Survey Type", "Plan Type", "Broadband RGU",
    "Customer Response Date (EST)", "Probabilidad de Recomendar",
    "Internet - Likelihood to Recommend", "Mobile - Likelihood to Recommend",
    *COMMENT_COLUMNS, "NPS Segments - pNPS Internet", "NPS Segments - pNPS Mobile",
    "NPS Segments - rNPS/tNPS",
]
SEGMENT_ORDER = [
    "pNPS Internet", "pNPS Mobile - Contrato", "pNPS Mobile - Prepago", "rNPS / Relación",
]
MAPPING = [
    ["Fuente", "Campo / regla", "Uso en reporte"],
    ["Fecha", "Customer Response Date (EST)", "Fecha de segmentación temporal; se normaliza a YYYY-MM-DD"],
    ["rNPS", "Survey Type = rNPS → Probabilidad de Recomendar", "Score 0–10 para NPS relacional"],
    ["pNPS Internet", "Survey Type = pNPS + Plan Type = Servicio residencial + Broadband RGU > 0 → Internet - Likelihood to Recommend", "Score 0–10 para pNPS Internet"],
    ["pNPS Mobile contrato", "Survey Type = pNPS + Plan Type contiene Contrato → Mobile - Likelihood to Recommend", "Score 0–10 para pNPS Mobile contrato"],
    ["pNPS Mobile prepago", "Survey Type = pNPS + Plan Type contiene Prepago → Mobile - Likelihood to Recommend", "Score 0–10 para pNPS Mobile prepago"],
    ["Clasificación", "9–10 Promotor; 7–8 Neutro; 0–6 Detractor", "Base estándar NPS"],
    ["NPS", "(% Promotores - % Detractores) × 100", "Calculado por segmento y mes"],
]


def text(value):
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat(sep=" ") if isinstance(value, datetime) else value.isoformat()
    return " ".join(str(value).split()).strip()


def as_float(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def nps(rows):
    if not rows:
        return 0.0
    promoters = sum(row["NPS Class"] == "Promotor" for row in rows)
    detractors = sum(row["NPS Class"] == "Detractor" for row in rows)
    return round((promoters - detractors) * 100 / len(rows), 1)


def classify_score(score):
    return "Promotor" if score >= 9 else "Neutro" if score >= 7 else "Detractor"


def product_and_score(get, survey_type, plan_type, broadband):
    survey = survey_type.lower()
    plan = plan_type.lower()
    if survey == "rnps":
        return "rNPS / Relación", get("Probabilidad de Recomendar"), "Probabilidad de Recomendar"
    if survey == "pnps" and plan == "servicio residencial" and broadband not in ("", "0"):
        return "pNPS Internet", get("Internet - Likelihood to Recommend"), "Internet - Likelihood to Recommend"
    if survey == "pnps" and "contrato" in plan:
        return "pNPS Mobile - Contrato", get("Mobile - Likelihood to Recommend"), "Mobile - Likelihood to Recommend"
    if survey == "pnps" and "prepago" in plan:
        return "pNPS Mobile - Prepago", get("Mobile - Likelihood to Recommend"), "Mobile - Likelihood to Recommend"
    return "No clasificado", None, ""


def discover_input(root):
    candidates = sorted(root.glob("CWP*.xlsx")) + sorted((root / "input").glob("CWP*.xlsx"))
    unique = list(dict.fromkeys(path.resolve() for path in candidates))
    if not unique:
        raise FileNotFoundError("No encontré un archivo CWP*.xlsx en la carpeta del proyecto ni en input\\.")
    if len(unique) > 1:
        names = ", ".join(path.name for path in unique)
        raise RuntimeError(f"Encontré varios archivos CWP*.xlsx: {names}. Usa --input para elegir uno.")
    return unique[0]


def comment_hash(comment):
    return hashlib.sha256(comment.encode("utf-8")).hexdigest()


def load_classification_state(state_path):
    if not state_path.exists():
        return {}
    try:
        payload = json.loads(state_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    if payload.get("version") != TAXONOMY_VERSION:
        return {}
    records = payload.get("records", {})
    return records if isinstance(records, dict) else {}


def write_classification_state(state_path, feedback_rows, previous_state=None):
    records = dict(previous_state or {})
    timestamp = datetime.now().isoformat(timespec="seconds")
    for row in feedback_rows:
        records[row["feedback_key"]] = {
            "comment_hash": row["comment_hash"],
            "category_auto": row["category_auto"],
            "category": row["category"],
            "category_source": row["category_source"],
            "updated_at": timestamp,
        }
    state_path.write_text(
        json.dumps({"version": TAXONOMY_VERSION, "records": records}, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )


def load_manual_overrides(review_path):
    if not review_path.exists():
        return {}
    overrides = {}
    with review_path.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            key = (row.get("feedback_key") or "").strip()
            category = (row.get("category") or "").strip()
            if key and category in TAXONOMY:
                source = (row.get("category_source") or "").strip()
                legacy_manual = category != (row.get("category_auto") or "").strip()
                explicit_source = source if source not in ("", "auto") else ""
                if explicit_source or legacy_manual:
                    overrides[key] = {"category": category, "source": explicit_source or "manual"}
    return overrides


def write_review_csv(review_path, feedback_rows):
    fields = [
        "feedback_key", "month", "segment", "nps_class", "score", "category_auto",
        "category", "category_source", "feedback",
    ]
    with review_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows({field: row.get(field, "") for field in fields} for row in feedback_rows)


def build_dataset(input_path, manual_overrides=None, classification_state=None):
    manual_overrides = manual_overrides or {}
    classification_state = classification_state or {}
    workbook = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    iterator = sheet.iter_rows(values_only=True)
    next(iterator, None)
    next(iterator, None)
    headers = list(next(iterator, ()))
    positions = {name: index for index, name in enumerate(headers) if name}
    missing = [name for name in NEEDED_COLUMNS if name not in positions]
    if missing:
        raise ValueError(f"Columnas faltantes en el Excel: {missing}")

    records, feedback_rows, raw_rows = [], [], 0
    auto_reused = auto_recalculated = manual_applied = cached_manual_reused = 0
    for row in iterator:
        raw_rows += 1

        def get(name):
            return text(row[positions[name]])

        unique_id = get("CW - Unique ID")
        if not unique_id:
            continue
        survey_type, plan_type, broadband = get("Survey Type"), get("Plan Type"), get("Broadband RGU")
        segment, score_raw, score_basis = product_and_score(get, survey_type, plan_type, broadband)
        score = as_float(score_raw)
        if score is None or not 0 <= score <= 10:
            continue
        response_date = get("Customer Response Date (EST)")[:10]
        nps_class = classify_score(score)
        record = {
            "CW - Unique ID": unique_id,
            "Unit": get("Unit"), "Survey Type": survey_type, "Plan Type": plan_type,
            "Broadband RGU": broadband, "Customer Response Date (EST)": get("Customer Response Date (EST)"),
            "Probabilidad de Recomendar": get("Probabilidad de Recomendar"),
            "Internet - Likelihood to Recommend": get("Internet - Likelihood to Recommend"),
            "Mobile - Likelihood to Recommend": get("Mobile - Likelihood to Recommend"),
            **{column: get(column) for column in COMMENT_COLUMNS},
            "NPS Segments - pNPS Internet": get("NPS Segments - pNPS Internet"),
            "NPS Segments - pNPS Mobile": get("NPS Segments - pNPS Mobile"),
            "NPS Segments - rNPS/tNPS": get("NPS Segments - rNPS/tNPS"),
            "Score": score, "Score Basis": score_basis, "Product Segment": segment,
            "NPS Class": nps_class, "Response Date": response_date,
        }
        records.append(record)
        comment = " | ".join(get(column) for column in COMMENT_COLUMNS if get(column))
        if comment:
            current_hash = comment_hash(comment)
            cached = classification_state.get(unique_id, {})
            cache_valid = (
                cached.get("comment_hash") == current_hash
                and cached.get("category_auto") in TAXONOMY
            )
            if cache_valid:
                category_auto = cached["category_auto"]
                auto_reused += 1
            else:
                category_auto = categorize(comment, segment)
                auto_recalculated += 1

            override = manual_overrides.get(unique_id)
            if override:
                category = override["category"]
                category_source = override["source"]
                manual_applied += 1
            elif cache_valid and cached.get("category_source") not in (None, "", "auto") and cached.get("category") in TAXONOMY:
                category = cached["category"]
                category_source = cached["category_source"]
                cached_manual_reused += 1
            else:
                category = category_auto
                category_source = "auto"
            feedback_rows.append({
                "feedback_key": unique_id, "month": response_date[:7], "segment": segment,
                "nps_class": nps_class, "category_auto": category_auto, "category": category,
                "category_source": category_source, "comment_hash": current_hash,
                "score": score, "feedback": comment,
                "category_local": category, "category_ollama": None,
            })
    workbook.close()

    def rows_for(segment):
        return records if segment == "Total" else [row for row in records if row["Product Segment"] == segment]

    segments = ["Total"] + [segment for segment in SEGMENT_ORDER if any(row["Product Segment"] == segment for row in records)]
    summary = []
    for segment in segments:
        rows = rows_for(segment)
        summary.append({
            "Segment": segment, "n": len(rows), "NPS": nps(rows),
            "Promoters": sum(row["NPS Class"] == "Promotor" for row in rows),
            "Neutrals": sum(row["NPS Class"] == "Neutro" for row in rows),
            "Detractors": sum(row["NPS Class"] == "Detractor" for row in rows),
            "Avg Score": round(sum(row["Score"] for row in rows) / len(rows), 2) if rows else 0,
        })

    monthly_groups = defaultdict(list)
    for row in records:
        monthly_groups[row["Response Date"][:7]].append(row)
    monthly = [{"Month": month, "n": len(rows), "NPS": nps(rows)} for month, rows in sorted(monthly_groups.items())]
    category_counts = Counter(row["category"] for row in feedback_rows)
    categories = [{"category": category, "count": count, "pct": round(count * 100 / max(len(feedback_rows), 1), 1), "by_segment": {}}
                  for category, count in category_counts.most_common()]

    return {
        "headers": headers, "needed": NEEDED_COLUMNS, "records": records, "summary": summary,
        "monthly": monthly, "sample": records[:10], "mapping": MAPPING,
        "source_file": str(input_path), "raw_rows": raw_rows,
        "report_updated_at": report_timestamp(),
        "feedback_model": {
            "model": "cx_drivers_v1_incremental_rules", "feedback_with_text": len(feedback_rows),
            "categories": categories, "rows": feedback_rows, "ollama": None,
            "taxonomy": TAXONOMY, "classifier": TAXONOMY_VERSION,
            "classification_note": "Clasificación local incremental: reutiliza comentarios sin cambios y conserva recategorizaciones editadas en feedback_review.csv.",
            "classification_stats": {
                "auto_reused": auto_reused,
                "auto_recalculated": auto_recalculated,
                "manual_or_external_applied": manual_applied,
                "cached_manual_or_external_reused": cached_manual_reused,
            },
        },
    }


def main():
    root = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(description="Genera el dashboard CX NPS desde un export CWP*.xlsx")
    parser.add_argument("--input", type=Path, help="Ruta al Excel CWP; si se omite, busca CWP*.xlsx en el proyecto")
    parser.add_argument("--output-dir", type=Path, default=root / "outputs", help="Carpeta local de resultados")
    parser.add_argument("--node", default="node", help="Ejecutable de Node.js")
    args = parser.parse_args()
    input_path = (args.input or discover_input(root)).resolve()
    if not input_path.exists():
        raise FileNotFoundError(f"No existe el Excel: {input_path}")
    output_dir = args.output_dir if args.output_dir.is_absolute() else root / args.output_dir
    output_dir.mkdir(parents=True, exist_ok=True)
    data_path, dashboard_path = output_dir / "nps_data.json", output_dir / "cx_nps_dashboard.html"
    review_path = output_dir / "feedback_review.csv"
    state_path = output_dir / "classification_state.json"
    state = load_classification_state(state_path)
    data = build_dataset(input_path, load_manual_overrides(review_path), state)
    data_path.write_text(json.dumps(data, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    write_review_csv(review_path, data["feedback_model"]["rows"])
    build_script = root / "work" / "build_dashboard_clean.mjs"
    subprocess.run([args.node, str(build_script), str(data_path), str(dashboard_path)], cwd=root, check=True)
    write_classification_state(state_path, data["feedback_model"]["rows"], state)
    print(json.dumps({"input": str(input_path), "raw_rows": data["raw_rows"], "valid_responses": len(data["records"]), "feedback_with_text": data["feedback_model"]["feedback_with_text"], "classification_stats": data["feedback_model"]["classification_stats"], "dashboard": str(dashboard_path), "data": str(data_path), "review": str(review_path), "state": str(state_path)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    try:
        main()
    except (FileNotFoundError, RuntimeError, ValueError) as error:
        print(f"Error: {error}", file=sys.stderr)
        raise SystemExit(2)
