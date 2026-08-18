"""Create the daily cumulative pNPS workbook from normalized CX NPS records."""

from calendar import monthrange
from collections import defaultdict
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


REPORT_SEGMENTS = (
    ("Fijo", "pNPS Internet"),
    ("Contrato", "pNPS Mobile - Contrato"),
    ("Prepago", "pNPS Mobile - Prepago"),
)

FULL_SPANISH_MONTHS = (
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
)

HEADERS = (
    "Día", "Muestras", "% Prom.", "% Neut.", "% Detr.", "NPS Diario",
    "Prom. (cant)", "Neut. (cant)", "Detr. (cant)", "Acum. Prom.",
    "Acum. Neut.", "Acum. Detr.", "Acum. Muestras", "NPS Acumulado",
)

HEADER_FILL = PatternFill("solid", fgColor="2F5496")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
NPS_FILL = PatternFill("solid", fgColor="DDEBF7")
WHITE_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=11)
TITLE_FONT = Font(name="Calibri", size=12, bold=True, color="1F1F1F")
THIN_BLUE = Side(style="thin", color="9EADBE")
GRID_BORDER = Border(left=THIN_BLUE, right=THIN_BLUE, top=THIN_BLUE, bottom=THIN_BLUE)


def _month_label(month_key):
    year, month = (int(value) for value in month_key.split("-"))
    return f"{FULL_SPANISH_MONTHS[month - 1]} {year}"


def _count_rows(rows):
    return {
        "samples": len(rows),
        "promoters": sum(row["NPS Class"] == "Promotor" for row in rows),
        "neutrals": sum(row["NPS Class"] == "Neutro" for row in rows),
        "detractors": sum(row["NPS Class"] == "Detractor" for row in rows),
    }


def _configure_sheet(sheet):
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A4"
    widths = (12, 13, 12, 12, 12, 14, 14, 14, 14, 15, 15, 15, 17, 18)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width


def _write_segment_block(sheet, start_row, month_key, segment_label, daily_rows, days_to_show):
    sheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(HEADERS))
    title = sheet.cell(start_row, 1, f"{segment_label} · NPS acumulado diario · {_month_label(month_key)}")
    title.fill = SECTION_FILL
    title.font = TITLE_FONT
    title.alignment = Alignment(horizontal="left", vertical="center")
    title.border = GRID_BORDER
    sheet.row_dimensions[start_row].height = 22

    header_row = start_row + 1
    for column, header in enumerate(HEADERS, start=1):
        cell = sheet.cell(header_row, column, header)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = GRID_BORDER
    sheet.row_dimensions[header_row].height = 30

    first_data_row = header_row + 1
    for offset, day in enumerate(range(1, days_to_show + 1)):
        row_number = first_data_row + offset
        counts = _count_rows(daily_rows.get(day, []))
        values = (
            day, counts["samples"], None, None, None, None,
            counts["promoters"], counts["neutrals"], counts["detractors"],
            None, None, None, None, None,
        )
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row_number, column, value)
            cell.font = BODY_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = GRID_BORDER

        sheet.cell(row_number, 3, f'=IF(B{row_number}=0,"",G{row_number}/B{row_number})')
        sheet.cell(row_number, 4, f'=IF(B{row_number}=0,"",H{row_number}/B{row_number})')
        sheet.cell(row_number, 5, f'=IF(B{row_number}=0,"",I{row_number}/B{row_number})')
        sheet.cell(row_number, 6, f'=IF(B{row_number}=0,"",(G{row_number}-I{row_number})/B{row_number}*100)')
        sheet.cell(row_number, 10, f'=SUM(G{first_data_row}:G{row_number})')
        sheet.cell(row_number, 11, f'=SUM(H{first_data_row}:H{row_number})')
        sheet.cell(row_number, 12, f'=SUM(I{first_data_row}:I{row_number})')
        sheet.cell(row_number, 13, f'=SUM(B{first_data_row}:B{row_number})')
        sheet.cell(row_number, 14, f'=IF(M{row_number}=0,"",(J{row_number}-L{row_number})/M{row_number}*100)')
        for column in range(3, 6):
            sheet.cell(row_number, column).number_format = "0.0%"
        for column in (6, 14):
            sheet.cell(row_number, column).number_format = "0.0"
            sheet.cell(row_number, column).fill = NPS_FILL
            sheet.cell(row_number, column).font = Font(name="Calibri", size=11, bold=True)
        sheet.row_dimensions[row_number].height = 19

    return first_data_row + days_to_show + 2


def write_daily_cumulative_workbook(records, output_path):
    """Write one month sheet with Fijo, Contrato and Prepago daily blocks."""
    grouped = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    relevant_segments = {segment for _, segment in REPORT_SEGMENTS}
    for record in records:
        if record["Product Segment"] not in relevant_segments:
            continue
        response_date = date.fromisoformat(record["Response Date"])
        grouped[response_date.strftime("%Y-%m")][record["Product Segment"]][response_date.day].append(record)

    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.title = "NPS acumulado diario"
    workbook.properties.subject = "Fijo, Contrato y Prepago"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    for month_key in sorted(grouped, reverse=True):
        year, month = (int(value) for value in month_key.split("-"))
        days_to_show = max(
            (day for segment_days in grouped[month_key].values() for day in segment_days),
            default=monthrange(year, month)[1],
        )
        sheet = workbook.create_sheet(_month_label(month_key))
        _configure_sheet(sheet)
        next_row = 1
        for segment_label, segment in REPORT_SEGMENTS:
            next_row = _write_segment_block(
                sheet, next_row, month_key, segment_label,
                grouped[month_key].get(segment, {}), days_to_show,
            )

    workbook.save(output_path)
    return {"months": len(grouped), "segments": [label for label, _ in REPORT_SEGMENTS]}
